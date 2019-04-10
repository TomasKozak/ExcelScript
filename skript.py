###############################################################################
##                                                                           ##
##     Idea behind is to make simple automation using openpyxl library       ##
##     which enables python to work with .xlsx files, pprint library         ##
##     wich enables to write python data strutctures to a text files         ##
##                                                                           ##
##     data of census tracts for different state county in America are       ##
##     in used .xlsx, task is to create .py file with dictionary that will   ##
##     have single states as key and population by census as value           ##
##                                                                           ##                                                                           ##
###############################################################################
import openpyxl, pprint


class CountyWorkbook:
    def __init__(self, xlsx_file, sheet_name):
        print('Opening workbook...')
        self.wb = openpyxl.load_workbook(xlsx_file)
        self.sheet = self.wb.get_sheet_by_name(sheet_name)
        self.countyData = {}


    def collectData(self):
        print('Reading ' + str(self.sheet.max_row) + ' rows...')
        for row in range(2, self.sheet.max_row):
            state = self.sheet['B' + str(row)].value
            county = self.sheet['C' + str(row)].value
            pop = self.sheet['D' + str(row)].value

            ## inicialize dictionaries, otherwise values cannot be stored in them
            ## it does nothing if key already exists --> profit
            self.countyData.setdefault(state, {})
            self.countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

            ## do actual calculation
            self.countyData[state][county]['tracts'] += 1
            self.countyData[state][county]['pop'] += int(pop)

    def printResults(self):
        print('Writing results...')
        resultFile = open('census2010.py', 'w')
        resultFile.write('allData = ' + pprint.pformat(self.countyData))
        resultFile.close()
        print('Done.')

wb = CountyWorkbook('censuspopdata.xlsx', 'Population by Census Tract')
wb.collectData()
wb.printResults()
